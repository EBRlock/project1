import flet as ft
from typing import Optional, Dict, List, Union
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


class Pessoa:
    def __init__(self, nome: str, idade: int, sexo: str, cargo: str, abdominal: Optional[int] = None,
                 flexao: Optional[int] = None, corrida: Optional[int] = None):
        self.nome = nome
        self.idade = idade
        self.sexo = sexo
        self.cargo = cargo
        self.abdominal = abdominal
        self.flexao = flexao
        self.corrida = corrida

    def __repr__(self):
        return f"Pessoa(nome='{self.nome}', idade={self.idade}, sexo='{self.sexo}', cargo='{self.cargo}')"


# Constantes
BLUE_ACCENT = ft.colors.BLUE_ACCENT_700
WHITE = ft.colors.WHITE
BORDER_RADIUS = 10
TEXT_FIELD_BORDER_COLOR = ft.colors.BLUE_GREY_400

pessoas: List[Pessoa] = []
pessoa_selecionada: Optional[Pessoa] = None
dados_pessoa: Dict[str, Dict[str, Optional[int]]] = {}


def create_text_field(label: str, keyboard_type: Optional[ft.KeyboardType] = None) -> ft.TextField:
    """Cria um TextField com estilo padrão."""
    return ft.TextField(
        label=label,
        width=300,
        border_color=TEXT_FIELD_BORDER_COLOR,
        keyboard_type=keyboard_type,
    )


def create_elevated_button(text: str, on_click, width: int = 200) -> ft.ElevatedButton:
    """Cria um ElevatedButton com estilo padrão."""
    return ft.ElevatedButton(
        text,
        on_click=on_click,
        width=width,
        style=ft.ButtonStyle(
            bgcolor=BLUE_ACCENT,
            color=WHITE,
            shape=ft.RoundedRectangleBorder(radius=BORDER_RADIUS),
        ),
    )


def main(page: ft.Page):
    page.title = "TAF App"
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.padding = 20
    page.spacing = 10
    page.window_width = 800
    page.window_height = 900
    page.theme_mode = ft.ThemeMode.LIGHT

    # Tema
    def change_theme(e):
        page.theme_mode = (
            ft.ThemeMode.DARK if page.theme_mode == ft.ThemeMode.LIGHT else ft.ThemeMode.LIGHT
        )
        page.update()

    theme_switch = ft.Switch(
        value=False,
        on_change=change_theme,
        label="",  # Removido o texto
    )

    #
    # Página Inicial (Homepage)
    #
    def open_selecao_taf(e):
        page.go("/selecao_taf")

    home_view = ft.View(
        "/",
        [
            ft.Container(
                content=ft.Text(
                    "Bem-vindo ao TAF App!",
                    size=40,
                    weight=ft.FontWeight.BOLD,
                ),
                alignment=ft.alignment.center,
                padding=20,
            ),
            create_elevated_button("Iniciar", on_click=open_selecao_taf),
        ],
        vertical_alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        padding=ft.padding.all(0),
    )

    #
    # Página de Seleção de TAF
    #
    def open_cadastro(e):
        page.go("/cadastro")

    selecao_taf_view = ft.View(
        "/selecao_taf",
        [
            ft.AppBar(title=ft.Text("Selecione o Tipo de TAF"), actions=[theme_switch]),
            ft.Column(
                [
                    create_elevated_button(
                        "TAF Convencional",
                        on_click=open_cadastro,
                        width=300,
                    ),
                    create_elevated_button(
                        "TAF Especializado",
                        on_click=open_cadastro,  # Mudar para página de cadastro específica
                        width=300,
                    ),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                tight=True,
            ),
        ],
        vertical_alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        appbar=ft.AppBar(title=ft.Text("Selecione o Tipo de TAF"), center_title=True),
    )

    #
    # Página de Cadastro
    #
    nome_field = create_text_field("Nome")
    idade_field = create_text_field("Idade", keyboard_type=ft.KeyboardType.NUMBER)
    sexo_radio = ft.RadioGroup(
        content=ft.Row(
            [
                ft.Radio(value="masculino", label="Masculino"),
                ft.Radio(value="feminino", label="Feminino"),
            ],
            alignment=ft.MainAxisAlignment.CENTER,  # Centraliza os botões
        ),
    )
    cargo_field = create_text_field("Cargo")

    def cadastrar_pessoa(e):
        global pessoa_selecionada
        nome = nome_field.value
        idade = idade_field.value
        sexo = sexo_radio.value
        cargo = cargo_field.value

        if not all([nome, idade, sexo, cargo]):
            page.show_snack_bar(
                ft.SnackBar(ft.Text("Por favor, preencha todos os campos."), open=True)
            )
            return

        try:
            idade = int(idade)
        except ValueError:
            page.show_snack_bar(
                ft.SnackBar(ft.Text("Idade deve ser um número inteiro."), open=True)
            )
            return

        pessoa = None  # Inicializa pessoa aqui
        if pessoa_selecionada:  # se tiver alguem selecionado atualiza os dados
            nome_original = pessoa_selecionada.nome  # Guarda o nome original

            pessoa_selecionada.nome = nome
            pessoa_selecionada.idade = idade
            pessoa_selecionada.sexo = sexo
            pessoa_selecionada.cargo = cargo
            pessoa = pessoa_selecionada  # atribui para mostrar no print

            # Preservar os dados de abdominal, flexão e corrida
            if nome_original in dados_pessoa:
                dados_pessoa[pessoa_selecionada.nome] = dados_pessoa.pop(
                    nome_original)  # Atualiza a chave no dicionário dados_pessoa
                dados_pessoa[pessoa_selecionada.nome]["abdominal"] = pessoa_selecionada.abdominal
                dados_pessoa[pessoa_selecionada.nome]["flexao"] = pessoa_selecionada.flexao
                dados_pessoa[pessoa_selecionada.nome]["corrida"] = pessoa_selecionada.corrida


        else:
            pessoa = Pessoa(nome, idade, sexo, cargo, None, None,
                            None)  # atribui o valor criado a pessoa
            pessoas.append(pessoa)  # Se não tiver ninguem selecionado adiciona

        print(f"Pessoa cadastrada: {pessoa}")
        page.go("/lista")

        nome_field.value = ""
        idade_field.value = ""
        sexo_radio.value = None
        cargo_field.value = ""
        pessoa_selecionada = None  # reseta a seleção
        page.update()

    cadastro_view = ft.View(
        "/cadastro",
        [
            ft.AppBar(title=ft.Text("Cadastro de Pessoa"), actions=[theme_switch], center_title=True),
            ft.Column(
                [
                    nome_field,
                    idade_field,
                    sexo_radio,
                    cargo_field,
                    create_elevated_button("Cadastrar", on_click=cadastrar_pessoa),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        ],
        vertical_alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        appbar=ft.AppBar(title=ft.Text("Cadastro de Pessoa"), center_title=True),
    )

    #
    # Página de Lista
    #
    search_field = ft.TextField(
        label="Buscar",
        on_change=lambda e: atualizar_lista_pessoas(),
        width=150,
        border_color=ft.colors.BLUE_GREY_400,
    )

    def generate_data_table() -> ft.DataTable:
        data_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Idade")),
                ft.DataColumn(ft.Text("Sexo")),
                ft.DataColumn(ft.Text("Cargo")),
                ft.DataColumn(ft.Text("Abdominal")),
                ft.DataColumn(ft.Text("Flexão")),
                ft.DataColumn(ft.Text("Corrida")),
            ],
            rows=[
                ft.DataRow(
                    [
                        ft.DataCell(ft.Text(p.nome)),
                        ft.DataCell(ft.Text(str(p.idade))),
                        ft.DataCell(ft.Text(p.sexo)),
                        ft.DataCell(ft.Text(p.cargo)),
                        ft.DataCell(ft.Text(str(p.abdominal) or "")),
                        ft.DataCell(ft.Text(str(p.flexao) or "")),
                        ft.DataCell(ft.Text(str(p.corrida) or "")),
                    ]
                )
                for p in pessoas
            ],
        )
        return data_table

    # Função para exibir a tabela de dados e o botão de impressão
    def show_data_table(e):
        page.go("/dados_todos")  # Navega para a página dados_todos

    def export_to_excel(e):
        wb = Workbook()
        ws = wb.active
        colunas = ["Nome", "Idade", "Sexo", "Cargo", "Abdominal", "Flexão", "Corrida"]

        # Escreve os cabeçalhos
        for col_num, column_title in enumerate(colunas, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = column_title

        # Escreve os dados
        for row_num, pessoa in enumerate(pessoas, 2):
            ws[f"A{row_num}"] = pessoa.nome
            ws[f"B{row_num}"] = pessoa.idade
            ws[f"C{row_num}"] = pessoa.sexo
            ws[f"D{row_num}"] = pessoa.cargo
            ws[f"E{row_num}"] = pessoa.abdominal
            ws[f"F{row_num}"] = pessoa.flexao
            ws[f"G{row_num}"] = pessoa.corrida

        # Ajustar a largura das colunas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Salvar o arquivo
        file_path = "dados_taf.xlsx"
        wb.save(file_path)

        # Mostrar um diálogo de sucesso
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Exportação Concluída"),
            content=ft.Text(f"Dados exportados para {file_path}"),
            actions=[
                ft.TextButton("OK", on_click=lambda _: page.close_dialog()),
            ],
        )
        page.dialog = dlg
        dlg.open = True
        page.update()

    def close_dlg():
        page.dialog.open = False  # Fecha o diálogo
        page.update()

    def atualizar_lista_pessoas():
        list_view.controls = [
            ft.AppBar(
                title=ft.Text("Lista de Pessoas"),
                actions=[theme_switch],
                center_title=True,
            ),
            ft.Row(
                [  # Adicionando os iconButtons
                    ft.IconButton(
                        icon=ft.icons.ADD,
                        tooltip="Adicionar Candidato",
                        on_click=open_cadastro,
                        style=ft.ButtonStyle(
                            bgcolor=ft.colors.BLUE_ACCENT_700,
                            color=ft.colors.WHITE,
                            shape=ft.RoundedRectangleBorder(radius=10),
                        ),
                    ),
                    search_field,
                    create_elevated_button(
                        "Ver Dados",
                        on_click=show_data_table,  # chama a função show_data_table
                        width=200,
                    ),
                ],
                alignment=ft.MainAxisAlignment.CENTER,
            ),

        ]

        search_term = search_field.value.lower()
        for pessoa in pessoas:
            if search_term in pessoa.nome.lower() or search_term in pessoa.cargo.lower():
                list_view.controls.append(
                    ft.Row(
                        [  # Usando Row para alinhar os botões
                            ft.ElevatedButton(
                                text=pessoa.nome,
                                on_click=lambda e, p=pessoa: selecionar_pessoa(p),
                                width=200,  # Aumentei o tamanho do butão
                                style=ft.ButtonStyle(
                                    bgcolor=ft.colors.BLUE_ACCENT_700,
                                    color=ft.colors.WHITE,
                                    padding=10,
                                    shape=ft.RoundedRectangleBorder(radius=10),
                                ),
                            ),
                            ft.IconButton(
                                icon=ft.icons.EDIT,  # Ícone de edição
                                tooltip="Editar",
                                on_click=lambda e, p=pessoa: editar_pessoa(
                                    p),  # Passa o objeto pessoa
                                icon_color=ft.colors.GREEN_500,
                            ),
                            ft.IconButton(
                                icon=ft.icons.DELETE,  # Ícone de exclusão
                                tooltip="Excluir",
                                on_click=lambda e, p=pessoa: excluir_pessoa(
                                    p),  # Passa o objeto pessoa
                                icon_color=ft.colors.RED_500,
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,  # Alinha os botões na horizontal
                    )
                )
        page.update()

    def excluir_pessoa(pessoa: Pessoa):
        pessoas.remove(pessoa)
        # Remove também do dicionário dados_pessoa
        if pessoa.nome in dados_pessoa:
            del dados_pessoa[pessoa.nome]
        atualizar_lista_pessoas()
        page.update()

    def editar_pessoa(pessoa: Pessoa):
        global pessoa_selecionada
        pessoa_selecionada = pessoa
        nome_field.value = pessoa.nome
        idade_field.value = str(pessoa.idade)
        sexo_radio.value = pessoa.sexo
        cargo_field.value = pessoa.cargo
        page.go("/cadastro")
        page.update()

    list_view = ft.View(
        "/lista",
        [
            ft.AppBar(
                title=ft.Text("Lista de Pessoas"),
                actions=[theme_switch],
                center_title=True,
            ),
            ft.Row(
                [  # Adicionando os iconButtons
                    ft.IconButton(
                        icon=ft.icons.ADD,
                        tooltip="Adicionar Candidato",
                        on_click=open_cadastro,
                        style=ft.ButtonStyle(
                            bgcolor=ft.colors.BLUE_ACCENT_700,
                            color=ft.colors.WHITE,
                            shape=ft.RoundedRectangleBorder(radius=10),
                        ),
                    ),
                    search_field,
                    create_elevated_button(
                        "Ver Dados",
                        on_click=show_data_table,  # Chama a função show_data_table
                        width=200,
                    ),
                ],
                alignment=ft.MainAxisAlignment.CENTER,
            ),
        ],
        vertical_alignment=ft.MainAxisAlignment.START,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        appbar=ft.AppBar(title=ft.Text("Lista de Pessoas"), center_title=True),
    )

    #
    # Página de Dados
    #
    abdominal_field = create_text_field("Abdominal")
    flexao_field = create_text_field("Flexão")
    corrida_field = create_text_field("Corrida")

    def salvar_dados(e):
        global pessoa_selecionada, dados_pessoa
        if pessoa_selecionada:
            try:
                abdominal = int(abdominal_field.value) if abdominal_field.value else None
                flexao = int(flexao_field.value) if flexao_field.value else None
                corrida = int(corrida_field.value) if corrida_field.value else None
            except ValueError:
                dialog = ft.AlertDialog(
                    modal=True,
                    title=ft.Text("Erro"),
                    content=ft.Text(f"Os campos Abdominal, Flexão e Corrida devem ser números inteiros."),
                    actions=[
                        ft.TextButton("OK", on_click=lambda _: page.close_dialog()),
                    ],
                )
                page.dialog = dialog
                dialog.open = True
                page.update()

                return

            dados_pessoa[pessoa_selecionada.nome] = {
                "abdominal": abdominal,
                "flexao": flexao,
                "corrida": corrida,
            }

            pessoa_selecionada.abdominal = abdominal
            pessoa_selecionada.flexao = flexao
            pessoa_selecionada.corrida = corrida

            print(
                f"Dados salvos para {pessoa_selecionada.nome}: {dados_pessoa[pessoa_selecionada.nome]}"
            )

            page.go("/lista")
            abdominal_field.value = ""
            flexao_field.value = ""
            corrida_field.value = ""
            page.update()

    def carregar_dados_existentes():
        global pessoa_selecionada, dados_pessoa
        if pessoa_selecionada:
            if pessoa_selecionada.nome in dados_pessoa:
                dados = dados_pessoa[pessoa_selecionada.nome]
                abdominal_field.value = (
                    str(dados.get("abdominal", ""))
                    if dados.get("abdominal") is not None
                    else ""
                )
                flexao_field.value = (
                    str(dados.get("flexao", "")) if dados.get("flexao") is not None else ""
                )
                corrida_field.value = (
                    str(dados.get("corrida", "")) if dados.get("corrida") is not None else ""
                )
            else:
                abdominal_field.value = ""
                flexao_field.value = ""
                corrida_field.value = ""
        page.update()

    dados_view = ft.View(
        "/dados",
        [
            ft.Column(
                [
                    abdominal_field,
                    flexao_field,
                    corrida_field,
                    create_elevated_button("Salvar", on_click=salvar_dados),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        ],
        vertical_alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        appbar=ft.AppBar(title=ft.Text("Dados Físicos"), actions=[theme_switch], center_title=True),
    )

    def selecionar_pessoa(pessoa: Pessoa):
        global pessoa_selecionada
        pessoa_selecionada = pessoa
        print(f"Pessoa selecionada: {pessoa_selecionada}")
        carregar_dados_existentes()
        page.go("/dados")

    #
    # Página de Dados Todos
    #
    def dados_todos_view():
        # Função interna para usar o contexto da página

        data_table = generate_data_table()

        return ft.View(
            route="/dados_todos",
            controls=[
                ft.AppBar(title=ft.Text("Dados dos Candidatos"), actions=[theme_switch], center_title=True),
                ft.Column(
                    [data_table],
                    scroll=ft.ScrollMode.ALWAYS,
                    width=600,
                ),
                create_elevated_button(
                    "Exportar para Excel",
                    on_click=export_to_excel,
                    width=200,
                ),
            ],
            vertical_alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        )

    #
    # Rotas
    #
    def route_change(route):
        page.views.clear()
        page.views.append(home_view)
        if page.route == "/selecao_taf":
            page.views.append(selecao_taf_view)
        elif page.route == "/cadastro":
            page.views.append(cadastro_view)
        elif page.route == "/lista":
            atualizar_lista_pessoas()
            page.views.append(list_view)
        elif page.route == "/dados":
            carregar_dados_existentes()
            page.views.append(dados_view)
        elif page.route == "/dados_todos":
            page.views.append(dados_todos_view())  # Adicionei aqui
        page.update()

    def view_pop(view):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)


if __name__ == "__main__":
    ft.app(target=main)